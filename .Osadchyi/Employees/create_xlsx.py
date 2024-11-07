import csv
from openpyxl import Workbook
from datetime import datetime

def calculate_age(birth_date):
    return datetime.now().year - birth_date.year

try:
    with open('employees.csv', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        headers = next(reader)

        workbook = Workbook()
        all_sheet = workbook.active
        all_sheet.title = "all"
        all_sheet.append(["№", *headers, "Вік"])

        # Створення аркушів для вікових категорій
        sheets = {
            "younger_18": workbook.create_sheet("younger_18"),
            "18-45": workbook.create_sheet("18-45"),
            "45-70": workbook.create_sheet("45-70"),
            "older_70": workbook.create_sheet("older_70")
        }

        # Додавання заголовків для кожного аркуша
        for sheet in sheets.values():
            sheet.append(["№", "Прізвище", "Ім'я", "По батькові", "Дата народження", "Вік"])

        # Запис даних в відповідні аркуші
        for idx, row in enumerate(reader, start=1):
            birth_date = datetime.strptime(row[4], '%Y-%m-%d')
            age = calculate_age(birth_date)
            row_with_age = [idx, *row, age]
            all_sheet.append(row_with_age)

            # Розподіл по вікових категоріях
            if age < 18:
                sheets["younger_18"].append([idx, *row[:3], row[4], age])
            elif 18 <= age <= 45:
                sheets["18-45"].append([idx, *row[:3], row[4], age])
            elif 45 < age <= 70:
                sheets["45-70"].append([idx, *row[:3], row[4], age])
            else:
                sheets["older_70"].append([idx, *row[:3], row[4], age])

    workbook.save("employees.xlsx")
    print("Ok")
except FileNotFoundError:
    print("Помилка: Файл 'employees.csv' не знайдено.")
except Exception as e:
    print(f"Помилка: {e}")
